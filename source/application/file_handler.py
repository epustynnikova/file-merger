import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from source.model.dto import ReadingInfo, SourceItem, SourceColumnData, InputFile, get_file_type, \
    FileTypeEnum, search

def _get_temp_path(original_path: str) -> str:
    directory = os.path.dirname(original_path)
    filename = os.path.basename(original_path)
    new_filename = f"temp_{filename}"
    return str(os.path.join(directory, new_filename))

class ExcelHandler:
    def __init__(self, input_file_path):
        self.df = None
        self.input_file = InputFile(input_file_path, get_file_type(input_file_path))

    def read_file(self) -> pd.DataFrame:
        self.df = pd.read_excel(self.input_file.path,
                                # dtype=str,
                                keep_default_na=False,
                                engine=self.input_file.type.open_lib)
        if self.input_file.type in [FileTypeEnum.XLSX, FileTypeEnum.XLS]:
            self.df.drop(self.df.columns[self.df.columns.str.contains('unnamed', case=False)], axis=1, inplace=True)
        return self.df

    def save_df(self, file_name=None) -> str:
        if file_name is None:
            return self._save(self.input_file.path)
        else:
            return self._save(file_name, delete_xlsb=False)

    def _save(self, file_name, delete_xlsb=True) -> str:
        if self.input_file.type in [FileTypeEnum.XLSX, FileTypeEnum.XLS]:
            return self._save_xls_xlsx(file_name)
        else:
            return self._save_xlsb(file_name, delete_xlsb)

    def _save_xlsb(self, file_name, delete_xlsb):
        wb = Workbook()
        ws = wb.active
        header_font = Font(
            name="Calibri",
            size=11,
            bold=False
        )
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"  # светло-зелёный как в файле
        )
        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        for col_num, column_name in enumerate(self.df.columns, start=1):
            cell = ws.cell(row=1, column=col_num, value=column_name)

            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        for row_num, row in enumerate(self.df.itertuples(index=False), start=2):
            for col_num, value in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_num, value=value)

        widths = [
            17.71, 15, 17.86, 17.86, 17.86, 17.86, 17.86, 17.86, 17.86,
            15, 15, 15, 15, 15, 15, 15, 15
        ]
        for col_idx, width in enumerate(widths, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        xslx_file_name = re.sub(r'\.xlsb$', '.xlsx', file_name)
        wb.save(xslx_file_name)

        if delete_xlsb:
            os.remove(self.input_file.path)
        return file_name

    def _save_xls_xlsx(self, file_name) -> str:
        wb = load_workbook(self.input_file.path, data_only=True)
        ws = wb.active
        for r_idx, row in self.df.iterrows():
            excel_row = r_idx + 2  # +2, т.к. pandas: 0 -> строка Excel 2 (если заголовок в строке 1)
            for c_idx, value in enumerate(row):
                excel_col = c_idx + 1
                cell = ws.cell(row=excel_row, column=excel_col)
                if cell.value != value:
                    cell.value = value  # стиль остаётся прежним
        temp_file_name = _get_temp_path(file_name)
        wb.save(temp_file_name)
        if os.path.exists(file_name):
            os.remove(file_name)
        os.rename(temp_file_name, file_name)
        return file_name


class SourceFileHandler(ExcelHandler):
    def __init__(self, reading_info: ReadingInfo, input_file):
        super().__init__(input_file)
        self.reading_info = reading_info

    def read_file(self) -> list[SourceItem]:
        super().read_file()
        source_items = []
        self.reading_info.columns_for_copy = [c for c in self.reading_info.columns_for_copy if
                                              c.column_name in self.df.columns and c.status_name in self.df.columns]
        for idx, row in self.df.iterrows():
            id_value = row[self.reading_info.id_column_name]
            columns_values = []
            for reading_column in self.reading_info.columns_for_copy:
                columns_values.append(SourceColumnData(
                    name=reading_column.column_name,
                    value=row[reading_column.column_name],
                    status=search(row[reading_column.status_name]),
                    status_name=reading_column.status_name
                ))
            source_items.append(SourceItem(
                id=id_value,
                id_name=row[self.reading_info.id_column_name],
                columns=sorted(columns_values, key=lambda c: c.name)))
        return source_items


class TargetFileHandler(ExcelHandler):
    def __init__(self, input_file):
        super().__init__(input_file)
